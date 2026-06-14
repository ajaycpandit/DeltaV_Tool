# ═══════════════════════════════════════════════════════════════════════════════
# RECIPE (BATCH_RECIPE / PROCEDURE) FHX PARSER + EXCEL BUILDER
# ───────────────────────────────────────────────────────────────────────────────
# A DeltaV procedure/recipe export (BATCH_RECIPE TYPE=PROCEDURE) is structurally
# different from Phase / Command EM / State EM exports:
#
#   BATCH_RECIPE                     ← the procedure container (header + formula params)
#     FORMULA_PARAMETER  (xN)        ← recipe-level parameters (the formula "columns")
#     UNIT_ALIAS                     ← unit binding + PFC graphics
#     PROCEDURE / PFC logic:
#       STEP            (xN)         ← each non-START step name *is* a Unit Procedure ref
#         STEP_PARAMETER (xN)        ← per-UP params, usually ORIGIN=DEFERRED → a formula param
#       TRANSITION      (xN)         ← gate expressions between steps
#       STEP_TRANSITION_CONNECTION   ← explicit graph edge  STEP → TRANSITION
#       TRANSITION_STEP_CONNECTION   ← explicit graph edge  TRANSITION → STEP
#     BATCH_RECIPE_FORMULA (xN)      ← named formulas (e.g. VACCINE H1) holding
#       ATTRIBUTE_INSTANCE → VALUE   ← the actual parameter values (CV / STRING_VALUE / SET)
#
# This module reuses the shared primitives from app.py (extract_block, sc, wf, wa,
# _safe_merge_and_write, and the style fills) — they are passed in via init().
# ═══════════════════════════════════════════════════════════════════════════════

import io, re, openpyxl
from openpyxl.utils import get_column_letter

# These are injected from app.py so we share one definition of each helper/style.
_H = {}   # helpers + styles container


def init(helpers):
    """Wire in shared helpers/styles from app.py. Call once at import time."""
    _H.update(helpers)


# ───────────────────────────────────────────────────────────────────────────────
# PARSER
# ───────────────────────────────────────────────────────────────────────────────

def _attr(block, pattern, default=''):
    m = re.search(pattern, block)
    return m.group(1).strip() if m else default


def _parse_value_block(ab):
    """Parse the inner VALUE { ... } of an ATTRIBUTE_INSTANCE into a normalized dict.

    Handles three shapes seen in recipe exports:
      analog   : VALUE { DESCRIPTION="" HIGH=100 LOW=0 SCALABLE=F CV=25 UNITS="°C" }
      discrete : VALUE { SET="S_CHR_COLUMN_SEL" STRING_VALUE="Not Used" CHANGEABLE=F }
      message  : VALUE { CV="some text" }
    """
    # CV may be numeric (CV=25) or quoted string (CV="text")
    cv_q = re.search(r'\bCV="((?:[^"\\]|\\.)*)"', ab)
    cv_n = re.search(r'\bCV=([^\s"}]+)', ab)
    sv   = re.search(r'STRING_VALUE="((?:[^"\\]|\\.)*)"', ab)
    st   = re.search(r'SET="([^"]*)"', ab)
    hi   = re.search(r'\bHIGH=([^\s}]+)', ab)
    lo   = re.search(r'\bLOW=([^\s}]+)', ab)
    un   = re.search(r'\bUNITS="([^"]*)"', ab)
    ch   = re.search(r'CHANGEABLE=([TF])', ab)

    if sv:                       # discrete / enumerated
        value, kind = sv.group(1), 'discrete'
    elif cv_q:                   # quoted CV (message/string)
        value, kind = cv_q.group(1), 'string'
    elif cv_n:                   # numeric CV (analog)
        value, kind = cv_n.group(1), 'analog'
    else:
        value, kind = '', 'empty'

    return {
        'value':      value,
        'kind':       kind,
        'set':        st.group(1) if st else '',
        'high':       hi.group(1) if hi else '',
        'low':        lo.group(1) if lo else '',
        'units':      un.group(1) if un else '',
        'changeable': (ch.group(1) == 'T') if ch else True,
    }


def parse_recipe_fhx(text):
    """Parse a BATCH_RECIPE procedure export into a structured dict."""
    extract_block = _H['extract_block']

    rm = re.search(r'BATCH_RECIPE\s+NAME="([^"]+)"([^\{]*)\{', text)
    if not rm:
        return None
    recipe_block = extract_block(text, rm.end() - 1)

    recipe = {
        'name':        rm.group(1),
        'type':        _attr(rm.group(2), r'TYPE=(\S+)'),
        'category':    _attr(rm.group(2), r'CATEGORY="([^"]+)"'),
        'description': _attr(recipe_block, r'DESCRIPTION="([^"]*)"'),
        'author':      _attr(recipe_block, r'AUTHOR="([^"]*)"'),
        'version':     _attr(recipe_block, r'VERSION="([^"]*)"'),
        'default_size':_attr(recipe_block, r'DEFAULT_BATCH_SIZE=(\S+)'),
        'min_size':    _attr(recipe_block, r'MINIMUM_BATCH_SIZE=(\S+)'),
        'max_size':    _attr(recipe_block, r'MAXIMUM_BATCH_SIZE=(\S+)'),
        'product_code':_attr(recipe_block, r'PRODUCT_CODE="([^"]*)"'),
    }

    # ── Unit alias(es) ──────────────────────────────────────────────────────────
    units = []
    for um in re.finditer(r'UNIT_ALIAS\s+NAME="([^"]+)"', recipe_block):
        ub = extract_block(recipe_block, recipe_block.index('{', um.end()))
        units.append({
            'alias':      um.group(1),
            'unit_class': _attr(ub, r'UNIT_CLASS="([^"]+)"'),
            'unit':       _attr(ub, r'UNIT="([^"]+)"'),
            'alias_type': _attr(ub, r'TYPE=(\S+)'),
        })
    recipe['units'] = units

    # ── Formula parameters (recipe-level parameter definitions) ─────────────────
    formula_params = []
    for fm in re.finditer(r'FORMULA_PARAMETER\s+NAME="([^"]+)"', recipe_block):
        fb = extract_block(recipe_block, recipe_block.index('{', fm.end()))
        formula_params.append({
            'name':        fm.group(1),
            'connection':  _attr(fb, r'CONNECTION=(\S+)'),
            'group':       _attr(fb, r'GROUP="([^"]*)"'),
            'description': _attr(fb, r'DESCRIPTION="([^"]*)"'),
            'locked':      _attr(fb, r'IS_PARAMETER_LOCKED=([TF])') == 'T',
        })
    recipe['formula_params'] = formula_params

    # ── Default values (procedure-level ATTRIBUTE_INSTANCE blocks) ──────────────
    # These sit directly under BATCH_RECIPE — one per parameter — and hold the
    # DEFAULT value plus HIGH/LOW/UNITS (analog) or SET/STRING_VALUE (discrete).
    # A named formula only stores OVERRIDES; anything not overridden inherits
    # the default defined here.
    defaults = {}
    # The 301 procedure-level defaults all precede the PFC/STEP section; the only
    # other ATTRIBUTE_INSTANCEs inside recipe_block are a few nested in STEPs.
    # Scope to the pre-step slice so we capture exactly the defaults.
    _step_start = recipe_block.find('STEP NAME=')
    _defaults_scope = recipe_block[:_step_start] if _step_start != -1 else recipe_block
    for am in re.finditer(r'ATTRIBUTE_INSTANCE\s+NAME="([^"]+)"\s*\{', _defaults_scope):
        ab = extract_block(_defaults_scope, am.end() - 1)
        defaults[am.group(1)] = _parse_value_block(ab)
    recipe['defaults'] = defaults

    # ── Steps (= Unit Procedure references) + their step parameters ─────────────
    steps = {}
    for sm in re.finditer(r'STEP\s+NAME="([^"]+)"[^\{]*\{', recipe_block):
        sb = extract_block(recipe_block, sm.end() - 1)
        rc = re.search(r'RECTANGLE=\s*\{\s*X=(-?\d+)\s*Y=(-?\d+)', sb)
        step_params = []
        for pm in re.finditer(r'STEP_PARAMETER\s+NAME="([^"]+)"\s*\{', sb):
            pb = extract_block(sb, pm.end() - 1)
            step_params.append({
                'name':        pm.group(1),
                'origin':      _attr(pb, r'ORIGIN=(\S+)'),
                'deferred_to': _attr(pb, r'DEFERRED_TO="([^"]*)"'),
                'group':       _attr(pb, r'GROUP="([^"]*)"'),
                'value':       _attr(pb, r'STRING_VALUE="([^"]*)"') or _attr(pb, r'\bCV=([^\s}]+)'),
            })
        steps[sm.group(1)] = {
            'description': _attr(sb, r'DESCRIPTION="([^"]*)"'),
            'x': int(rc.group(1)) if rc else 0,
            'y': int(rc.group(2)) if rc else 0,
            'is_up': sm.group(1).upper() != 'START',
            'acquire_unit': _attr(sb, r'ACQUIRE_UNIT=([TF])') == 'T',
            'params': step_params,
        }

    # ── Transitions ─────────────────────────────────────────────────────────────
    trans = {}
    for tm in re.finditer(r'TRANSITION\s+NAME="([^"]+)"\s*\{', recipe_block):
        tb = extract_block(recipe_block, tm.end() - 1)
        pos = re.search(r'POSITION=\s*\{\s*X=(-?\d+)\s*Y=(-?\d+)', tb)
        trans[tm.group(1)] = {
            'description': _attr(tb, r'DESCRIPTION="([^"]*)"'),
            'expression':  _attr(tb, r'EXPRESSION="(.*?)"\s*\}', '') or
                           _attr(tb, r'EXPRESSION="(.*)"'),
            'termination': _attr(tb, r'TERMINATION=(\w+)', 'F'),
            'x': int(pos.group(1)) if pos else 0,
            'y': int(pos.group(2)) if pos else 0,
        }
    # EXPRESSION can be multi-line; re-grab with DOTALL for safety
    for tm in re.finditer(r'TRANSITION\s+NAME="([^"]+)"\s*\{', recipe_block):
        tb = extract_block(recipe_block, tm.end() - 1)
        em = re.search(r'EXPRESSION="(.*?)"(?=\s*\})', tb, re.DOTALL)
        if em:
            trans[tm.group(1)]['expression'] = em.group(1).strip()

    # ── Connection graph (explicit edges) ───────────────────────────────────────
    s2t, t2s = {}, {}
    for cm in re.finditer(r'STEP_TRANSITION_CONNECTION\s+STEP="([^"]+)"\s+TRANSITION="([^"]+)"', recipe_block):
        s2t.setdefault(cm.group(1), []).append(cm.group(2))
    for cm in re.finditer(r'TRANSITION_STEP_CONNECTION\s+TRANSITION="([^"]+)"\s+STEP="([^"]+)"', recipe_block):
        t2s.setdefault(cm.group(1), []).append(cm.group(2))

    recipe['steps'] = steps
    recipe['ordered_steps'] = sorted(steps.items(), key=lambda kv: kv[1]['y'])
    recipe['transitions'] = trans
    recipe['step_to_trans'] = s2t
    recipe['trans_to_step'] = t2s

    # ── Formulas (named value sets) ─────────────────────────────────────────────
    # NOTE: BATCH_RECIPE_FORMULA blocks are siblings of BATCH_RECIPE at the top
    # level of the file, NOT nested inside it — so scan the full text.
    formulas = []
    for bm in re.finditer(r'BATCH_RECIPE_FORMULA\s+NAME="([^"]+)"[^\{]*\{', text):
        bb = extract_block(text, bm.end() - 1)
        values = {}
        for am in re.finditer(r'ATTRIBUTE_INSTANCE\s+NAME="([^"]+)"\s*\{', bb):
            ab = extract_block(bb, am.end() - 1)
            values[am.group(1)] = _parse_value_block(ab)
        formulas.append({
            'name':        bm.group(1),
            'description': _attr(bb, r'DESCRIPTION="([^"]*)"'),
            'version':     _attr(bb, r'VERSION="([^"]*)"'),
            'released':    _attr(bb, r'RELEASED_TO_PRODUCTION=([TF])') == 'T',
            'values':      values,
        })
    recipe['formulas'] = formulas

    return recipe


# ───────────────────────────────────────────────────────────────────────────────
# EXCEL BUILDER
# ───────────────────────────────────────────────────────────────────────────────

def build_recipe_excel(recipe, fname, opts):
    """Build a styled multi-sheet workbook for a parsed recipe."""
    sc   = _H['sc']
    wf   = _H['wf']
    wa   = _H['wa']
    smaw = _H['_safe_merge_and_write']
    Alignment = _H['Alignment']
    S = _H['styles']  # dict of PatternFill/Border styles

    wb = openpyxl.Workbook()

    # ═══ SHEET 1: OVERVIEW ══════════════════════════════════════════════════════
    ws = wb.active
    ws.title = 'OVERVIEW'
    for ci, w in enumerate([22, 40, 22, 40], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    smaw(ws, 1, 1, 1, 4,
         value=f"  {recipe['name']}   —   {recipe['type']} Recipe",
         font=wf(True, 14, 'FFFFFF'), fill=S['NAVY'],
         alignment=Alignment(horizontal='left', vertical='center'), border=S['BORD'])
    ws.row_dimensions[1].height = 30
    smaw(ws, 2, 1, 2, 4, value=f"  {recipe['description']}",
         font=wf(False, 10, 'CCCCCC'), fill=S['NAVY'],
         alignment=Alignment(horizontal='left', vertical='center'), border=S['BORD'])
    ws.row_dimensions[2].height = 18

    info = [
        ('Author', recipe['author'], 'Version', recipe['version']),
        ('Category', recipe['category'], 'Product Code', recipe['product_code'] or '—'),
        ('Default Batch Size', recipe['default_size'], 'Min / Max',
         f"{recipe['min_size']} / {recipe['max_size']}"),
    ]
    r = 4
    for a, b, c, d in info:
        sc(ws, r, 1, a, bold=True, fill=S['BLUE_H'], fc='FFFFFF')
        sc(ws, r, 2, b, fill=S['WHITE'])
        sc(ws, r, 3, c, bold=True, fill=S['BLUE_H'], fc='FFFFFF')
        sc(ws, r, 4, d, fill=S['WHITE'])
        ws.row_dimensions[r].height = 16
        r += 1

    # Units
    r += 1
    smaw(ws, r, 1, r, 4, value='  UNIT ALIASES',
         font=wf(True, 10, 'FFFFFF'), fill=S['BLUE_H'],
         alignment=Alignment(horizontal='left', vertical='center'), border=S['BORD'])
    ws.row_dimensions[r].height = 16
    r += 1
    for h, ci in [('Alias', 1), ('Unit Class', 2), ('Bound Unit', 3), ('Type', 4)]:
        sc(ws, r, ci, h, bold=True, fill=S['BLUE_S'], fc='FFFFFF', h='center')
    r += 1
    for i, u in enumerate(recipe['units']):
        f = S['ALT'] if i % 2 == 0 else S['WHITE']
        sc(ws, r, 1, u['alias'], bold=True, fill=f)
        sc(ws, r, 2, u['unit_class'], fill=f)
        sc(ws, r, 3, u['unit'], fill=f)
        sc(ws, r, 4, u['alias_type'], fill=f, h='center')
        r += 1

    # Counts strip
    r += 1
    counts = [
        ('Unit Procedures', sum(1 for _, s in recipe['ordered_steps'] if s['is_up'])),
        ('Transitions', len(recipe['transitions'])),
        ('Formula Parameters', len(recipe['formula_params'])),
        ('Formulas', len(recipe['formulas'])),
    ]
    for label, n in counts:
        sc(ws, r, 1, label, bold=True, fill=S['BLUE_H'], fc='FFFFFF')
        sc(ws, r, 2, n, fill=S['WHITE'], h='center')
        ws.row_dimensions[r].height = 16
        r += 1
    ws.freeze_panes = 'A3'

    # ═══ SHEET 2: PROCEDURE (UP sequence + transitions) ═════════════════════════
    if opts.get('procedure', True):
        ws = wb.create_sheet('PROCEDURE')
        for ci, w in enumerate([8, 34, 44, 64], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        smaw(ws, 1, 1, 1, 4,
             value=f"  Procedural Logic   —   {recipe['name']}",
             font=wf(True, 13, 'FFFFFF'), fill=S['NAVY'],
             alignment=Alignment(horizontal='left', vertical='center'), border=S['BORD'])
        ws.row_dimensions[1].height = 26
        for h, ci in [('Seq', 1), ('Unit Procedure / Step', 2), ('Description', 3),
                      ('Outgoing Transition(s)', 4)]:
            sc(ws, 2, ci, h, bold=True, fill=S['BLUE_H'], fc='FFFFFF', h='center')
        ws.row_dimensions[2].height = 17
        ws.freeze_panes = 'A3'
        r = 3
        seq = 0
        for sn, sd in recipe['ordered_steps']:
            f = S['ALT'] if seq % 2 == 0 else S['WHITE']
            if sd['is_up']:
                seq += 1
                seqlabel = str(seq)
            else:
                seqlabel = '▸'
            outs = recipe['step_to_trans'].get(sn, [])
            out_desc = []
            for tn in outs:
                tr = recipe['transitions'].get(tn, {})
                nexts = recipe['trans_to_step'].get(tn, [])
                tag = '⏹' if tr.get('termination') == 'T' else '→'
                out_desc.append(f"{tag} {tn}" + (f"  ⇒ {', '.join(nexts)}" if nexts else ''))
            sc(ws, r, 1, seqlabel, fill=f, h='center', bold=True)
            sc(ws, r, 2, sn, fill=f, bold=sd['is_up'],
               fc='0C447C' if sd['is_up'] else '888888')
            sc(ws, r, 3, sd['description'], fill=f, wrap=True)
            sc(ws, r, 4, '\n'.join(out_desc), fill=f, wrap=True)
            ws.row_dimensions[r].height = max(16, 14 * max(1, len(out_desc)))
            r += 1

        # Transition expressions block
        r += 1
        smaw(ws, r, 1, r, 4, value='  TRANSITION EXPRESSIONS',
             font=wf(True, 11, 'FFFFFF'), fill=S['GREEN_S'],
             alignment=Alignment(horizontal='left', vertical='center'), border=S['BORD'])
        ws.row_dimensions[r].height = 18
        r += 1
        for h, ci in [('Transition', 1), ('Description', 2), ('Type', 3), ('Expression', 4)]:
            sc(ws, r, ci, h, bold=True, fill=S['BLUE_H'], fc='FFFFFF', h='center')
        r += 1
        for i, (tn, tr) in enumerate(recipe['transitions'].items()):
            f = S['ALT_G'] if i % 2 == 0 else S['ALTG2']
            sc(ws, r, 1, tn, fill=f, bold=True)
            sc(ws, r, 2, tr['description'], fill=f, wrap=True)
            sc(ws, r, 3, '⏹ END' if tr['termination'] == 'T' else '→ NEXT', fill=f, h='center')
            sc(ws, r, 4, tr['expression'], fill=f, wrap=True)
            ws.row_dimensions[r].height = max(16, min(120, 14 * max(1, len(tr['expression']) // 60 + 1)))
            r += 1

    # ═══ SHEET 3: FORMULA PARAMETERS ════════════════════════════════════════════
    if opts.get('parameters', True):
        ws = wb.create_sheet('PARAMETERS')
        defaults = recipe.get('defaults', {})
        #            Param  Type  Default Units  Low   High  Set    Conn  Group Description
        widths = [46, 10, 16, 9, 9, 9, 22, 11, 14, 46]
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ncol = len(widths)
        smaw(ws, 1, 1, 1, ncol,
             value=f"  Formula Parameters   ({len(recipe['formula_params'])})",
             font=wf(True, 13, 'FFFFFF'), fill=S['NAVY'],
             alignment=Alignment(horizontal='left', vertical='center'), border=S['BORD'])
        ws.row_dimensions[1].height = 26
        headers = ['Parameter', 'Type', 'Default', 'Units', 'Low', 'High',
                   'Set (enumeration)', 'Connection', 'Group', 'Description']
        for ci, h in enumerate(headers, 1):
            sc(ws, 2, ci, h, bold=True, fill=S['BLUE_H'], fc='FFFFFF', h='center')
        ws.row_dimensions[2].height = 17
        ws.freeze_panes = 'B3'
        r = 3
        for i, p in enumerate(recipe['formula_params']):
            f = S['ALT'] if i % 2 == 0 else S['WHITE']
            d = defaults.get(p['name'], {})
            kind = d.get('kind', '')
            kind_disp = {'analog': 'Analog', 'discrete': 'Discrete',
                         'string': 'String', 'empty': '', '': ''}.get(kind, kind)
            sc(ws, r, 1, p['name'], fill=f, bold=True)
            sc(ws, r, 2, kind_disp, fill=f, h='center',
               fc='2c6b70' if kind == 'analog' else '7a3d2a' if kind == 'discrete' else '777777')
            sc(ws, r, 3, d.get('value', ''), fill=f, h='center',
               bold=not d.get('changeable', True))
            sc(ws, r, 4, d.get('units', ''), fill=f, h='center')
            sc(ws, r, 5, d.get('low', ''),  fill=f, h='center', fc='777777')
            sc(ws, r, 6, d.get('high', ''), fill=f, h='center', fc='777777')
            sc(ws, r, 7, d.get('set', ''),  fill=f, h='center', fc='555555')
            sc(ws, r, 8, p['connection'], fill=f, h='center')
            sc(ws, r, 9, p['group'], fill=f, h='center')
            sc(ws, r, 10, p['description'], fill=f, wrap=True)
            ws.row_dimensions[r].height = 15
            r += 1

    # ═══ SHEET 4: FORMULA VALUES (matrix: param × formula) ══════════════════════
    if opts.get('formulas', True) and recipe['formulas']:
        ws = wb.create_sheet('FORMULA VALUES')
        formulas = recipe['formulas']
        defaults = recipe.get('defaults', {})

        # Parameter ordering: formula-param definition order, then any extras that
        # appear only as defaults or overrides.
        ordered = [p['name'] for p in recipe['formula_params']]
        seen = set(ordered)
        for src in (defaults.keys(), *[fo['values'].keys() for fo in formulas]):
            for k in src:
                if k not in seen:
                    ordered.append(k); seen.add(k)

        # Default view: params that have a default OR any override.
        # all_params: every parameter in the union.
        if opts.get('all_params', False):
            active = ordered
        else:
            active = [p for p in ordered
                      if p in defaults or any(p in fo['values'] for fo in formulas)]

        show_lim = opts.get('show_limits', True)
        # Columns: Param | Units | Default | [Low | High] | <formula cols...>
        base_cols = ['Parameter', 'Units', 'Default']
        if show_lim:
            base_cols += ['Low', 'High']
        nbase = len(base_cols)
        ncols = nbase + len(formulas)

        ws.column_dimensions['A'].width = 46
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 16
        if show_lim:
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 10
        for ci in range(nbase + 1, ncols + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 16

        smaw(ws, 1, 1, 1, ncols,
             value=f"  Formula Values   —   {recipe['name']}"
                   f"   ·   empty = inherits default",
             font=wf(True, 13, 'FFFFFF'), fill=S['NAVY'],
             alignment=Alignment(horizontal='left', vertical='center'), border=S['BORD'])
        ws.row_dimensions[1].height = 26

        for ci, h in enumerate(base_cols, 1):
            sc(ws, 2, ci, h, bold=True, fill=S['BLUE_H'], fc='FFFFFF', h='center')
        for ci, fo in enumerate(formulas, nbase + 1):
            sc(ws, 2, ci, fo['name'], bold=True, fill=S['ORANGE_H'], fc='FFFFFF', h='center')
        ws.row_dimensions[2].height = 18
        # Description sub-row (only under formula cols)
        for ci in range(1, nbase + 1):
            sc(ws, 3, ci, '', fill=S['BLUE_H'])
        for ci, fo in enumerate(formulas, nbase + 1):
            sc(ws, 3, ci, fo['description'], fill=S['BLUE_S'], fc='FFFFFF', h='center', wrap=True)
        ws.row_dimensions[3].height = 16
        ws.freeze_panes = get_column_letter(nbase + 1) + '4'

        r = 4
        for i, pname in enumerate(active):
            f = S['ALT'] if i % 2 == 0 else S['WHITE']
            dft = defaults.get(pname, {})
            dft_disp = dft.get('value', '') if dft else ''
            sc(ws, r, 1, pname, fill=f, bold=True)
            sc(ws, r, 2, dft.get('units', ''), fill=f, h='center')
            sc(ws, r, 3, dft_disp, fill=f, h='center',
               bold=not dft.get('changeable', True))
            if show_lim:
                sc(ws, r, 4, dft.get('low', ''),  fill=f, h='center', fc='777777')
                sc(ws, r, 5, dft.get('high', ''), fill=f, h='center', fc='777777')
            for ci, fo in enumerate(formulas, nbase + 1):
                v = fo['values'].get(pname)
                if v is None:
                    # Inherits the default — show it greyed/italic so it's clearly
                    # not an explicit override.
                    sc(ws, r, ci, dft_disp if dft_disp != '' else '—',
                       fill=f, h='center', fc='B0B0B0')
                else:
                    sc(ws, r, ci, v['value'], fill=f, h='center',
                       bold=not v['changeable'], fc='0C447C')
            ws.row_dimensions[r].height = 15
            r += 1

    # ═══ SHEET 5 (toggle): STEP PARAMETERS — per-UP deferred-parameter links ════
    if opts.get('step_params', False):
        ws = wb.create_sheet('STEP PARAMETERS')
        for ci, w in enumerate([34, 40, 14, 46, 18], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        smaw(ws, 1, 1, 1, 5,
             value=f"  Step Parameters   —   per-UP parameter links",
             font=wf(True, 13, 'FFFFFF'), fill=S['NAVY'],
             alignment=Alignment(horizontal='left', vertical='center'), border=S['BORD'])
        ws.row_dimensions[1].height = 26
        for h, ci in [('Unit Procedure', 1), ('Step Parameter', 2), ('Origin', 3),
                      ('Deferred To (Formula Param)', 4), ('Group', 5)]:
            sc(ws, 2, ci, h, bold=True, fill=S['BLUE_H'], fc='FFFFFF', h='center')
        ws.row_dimensions[2].height = 17
        ws.freeze_panes = 'A3'
        r = 3
        shade = 0
        for sn, sd in recipe['ordered_steps']:
            if not sd['is_up'] or not sd['params']:
                continue
            # UP group header
            smaw(ws, r, 1, r, 5,
                 value=f"  {sn}   ({len(sd['params'])} parameter{'s' if len(sd['params']) != 1 else ''})",
                 font=wf(True, 10, 'FFFFFF'), fill=S['BLUE_S'],
                 alignment=Alignment(horizontal='left', vertical='center'), border=S['BORD'])
            ws.row_dimensions[r].height = 16
            r += 1
            for i, p in enumerate(sd['params']):
                f = S['ALT'] if i % 2 == 0 else S['WHITE']
                deferred = p['origin'].upper() == 'DEFERRED'
                sc(ws, r, 1, '', fill=f)
                sc(ws, r, 2, p['name'], fill=f, bold=True)
                sc(ws, r, 3, p['origin'], fill=f, h='center',
                   fc='0C447C' if deferred else '888888')
                sc(ws, r, 4, p['deferred_to'] if deferred else (p['value'] or '—'),
                   fill=f, wrap=True,
                   fc='000000' if deferred else '555555')
                sc(ws, r, 5, p['group'], fill=f, h='center')
                ws.row_dimensions[r].height = 15
                r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
